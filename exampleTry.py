import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

url = "https://apis.cbs.gov.il/series/data/list?id=62916&format=xls&download=false"

params = {
    "id": "6,2,9,1,6",  #path to data - id code
    "format": "xls",
    "download": "false"
}

response = requests.get(url, params=params)

root = ET.fromstring(response.content)
ind_elements = root.findall('./ind')

data_2022 = {}

#go through all <ind> elements
for ind in ind_elements:

    name = ind.find('name').text
    date = ind.find('date').text
    if date.startswith('2022'):
        index = ind.find('index').text
        percent = ind.find('percent').text
        code = ind.find('code').text
        base = ind.find('base').text

        data_2022[date] = {'name': name, 'index': index, 'percent': percent, 'code': code, 'base': base}

wb = load_workbook(filename='rdata_new_copy.xlsx')

ws = wb['quarterly']

start_row = 145

for date, data in data_2022.items():
    ws.cell(row=start_row, column=32, value=data['index'])
    start_row += 1

wb.save('rdata_new_copy.xlsx')
