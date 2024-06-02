"""import requests
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def get_data(id_code):
    url = "https://apis.cbs.gov.il/series/data/path?id=37&format=xls&download=true"

    params = {
        "id": id_code,  #path to data - id code
        "format": "xls",
        "download": "true"
    }

    response = requests.get(url, params=params)
    data = pd.read_excel(BytesIO(response.content), engine='openpyxl')

data_2022 = get_data(row['code_sidr'])
    last_row = ws.max_row
    for i, (date, data) in enumerate(data_2022.items()):
        #התאמה בין הטורי הסבר לכותרות העמודות בקובץ שלנו
        for top_code in ['top_code2', 'top_code3', 'top_code4', 'top_code5', 'top_code6']:
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value == row[top_code]:
                    column = col[0].column
                    ws.cell(row=last_row+i+1, column=column, value=data['index'])

df_codes = pd.read_excel('top_code 37 for ozar.xlsx')

wb = load_workbook(filename='Try.xlsx')

wb.save('Try.xlsx')
"""
import requests
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

def get_data(id_code):
    url = f"https://apis.cbs.gov.il/series/data/path"
    params = {
        "id": id_code,
        "format": "xls",
        "download": "true"
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = pd.read_excel(BytesIO(response.content), engine='openpyxl')
        return data[data['date'].str.startswith('2022-')]
    else:
        return pd.DataFrame()  # Returns an empty DataFrame if the request fails

# Load your Excel file with id codes and descriptions
df_codes = pd.read_excel('top_code 37 for ozar.xlsx')
wb = load_workbook(filename='Try.xlsx')
time_code_dict = {1: 'yearly', 2: 'quarterly', 3: 'monthly'}

for index, row in df_codes.iterrows():
    sheet_name = time_code_dict[row['time_code']]
    ws = wb[sheet_name]
    data_2022 = get_data(row['code_sidr'])
    last_row = ws.max_row
    for index, row_data in data_2022.iterrows():
        for top_code in ['top_code2', 'top_code3', 'top_code4', 'top_code5', 'top_code6']:
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value == row[top_code]:
                    ws.cell(row=last_row+index+1, column=col[0].column, value=row_data['index'])

wb.save('Try.xlsx')